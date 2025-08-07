
"""
Database export utility for Monthly Organics
Exports all database data to JSON, CSV, or XLSX format with optional decryption
"""
import json
import logging
import csv
import io
from datetime import datetime, date
from decimal import Decimal
from services.database import DatabaseService
from utils.encryption import SecureDataHandler, DataEncryption

logger = logging.getLogger(__name__)

class DatabaseExporter:
    """Handles exporting database data to JSON format"""
    
    @staticmethod
    def _decrypt_row_data(table_name, row_dict):
        """Decrypt encrypted fields in a row based on table type"""
        try:
            decrypted_row = row_dict.copy()
            
            # Handle user table encryption
            if table_name == 'users':
                if 'phone_encrypted' in row_dict:
                    decrypted_phone = DataEncryption.decrypt_phone(row_dict['phone_encrypted'])
                    if decrypted_phone:
                        decrypted_row['phone_decrypted'] = decrypted_phone
                    # Keep encrypted version for reference
                    
            # Handle address table encryption
            elif table_name == 'addresses':
                encrypted_fields = {
                    'house_number_encrypted': 'house_number_decrypted',
                    'floor_door_encrypted': 'floor_door_decrypted',
                    'contact_number_encrypted': 'contact_number_decrypted',
                    'nearby_landmark_encrypted': 'nearby_landmark_decrypted',
                    'receiver_name_encrypted': 'receiver_name_decrypted'
                }
                
                for encrypted_field, decrypted_field in encrypted_fields.items():
                    if encrypted_field in row_dict and row_dict[encrypted_field]:
                        decrypted_value = DataEncryption.decrypt_address_field(row_dict[encrypted_field])
                        if decrypted_value:
                            decrypted_row[decrypted_field] = decrypted_value
            
            return decrypted_row
            
        except Exception as e:
            logger.error(f"Error decrypting row data for table {table_name}: {e}")
            return row_dict
    
    @staticmethod
    def serialize_value(value, for_excel=False):
        """Convert database values to JSON-serializable format"""
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        elif isinstance(value, Decimal):
            return float(value)
        elif value is None:
            return None
        elif isinstance(value, dict) and for_excel:
            # Convert complex objects to strings for Excel compatibility
            return json.dumps(value)
        elif isinstance(value, (list, tuple)) and for_excel:
            # Convert arrays to strings for Excel compatibility
            return json.dumps(value)
        else:
            return value
    
    @staticmethod
    def get_table_data(table_name, for_excel=False, decrypt_data=False):
        """Get all data from a specific table with optional decryption"""
        try:
            # Check if table has an 'id' column for ordering
            has_id_query = f"""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = '{table_name}' AND column_name = 'id'
            """
            has_id = DatabaseService.execute_query(has_id_query)
            
            # Use appropriate ORDER BY clause
            if has_id:
                query = f"SELECT * FROM {table_name} ORDER BY id"
            else:
                query = f"SELECT * FROM {table_name}"
                
            result = DatabaseService.execute_query(query)
            
            if not result:
                return []
            
            # Convert to list of dictionaries with serialized values
            table_data = []
            for row in result:
                row_dict = dict(row)
                
                # Apply decryption if requested
                if decrypt_data:
                    row_dict = DatabaseExporter._decrypt_row_data(table_name, row_dict)
                
                # Serialize values for export
                processed_row = {}
                for key, value in row_dict.items():
                    processed_row[key] = DatabaseExporter.serialize_value(value, for_excel=for_excel)
                
                table_data.append(processed_row)
            
            return table_data
            
        except Exception as e:
            logger.error(f"Error exporting table {table_name}: {e}")
            return []
    
    @staticmethod
    def get_all_table_names():
        """Get list of all user tables in the database"""
        try:
            query = """
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_type = 'BASE TABLE'
                AND table_name != 'spatial_ref_sys'
                ORDER BY table_name
            """
            result = DatabaseService.execute_query(query)
            
            if not result:
                return []
            
            return [row['table_name'] for row in result]
            
        except Exception as e:
            logger.error(f"Error getting table names: {e}")
            return []
    
    @staticmethod
    def export_full_database(decrypt_data=False):
        """Export entire database to JSON format with optional decryption"""
        try:
            export_data = {
                'export_metadata': {
                    'export_timestamp': datetime.utcnow().isoformat(),
                    'database_name': 'monthly_organics',
                    'export_type': 'full_database',
                    'data_decrypted': decrypt_data
                },
                'tables': {}
            }
            
            # Get all table names
            table_names = DatabaseExporter.get_all_table_names()
            
            if not table_names:
                logger.warning("No tables found in database")
                return None
            
            # Export data from each table
            total_records = 0
            for table_name in table_names:
                logger.info(f"Exporting table: {table_name} (decrypt: {decrypt_data})")
                table_data = DatabaseExporter.get_table_data(table_name, decrypt_data=decrypt_data)
                export_data['tables'][table_name] = {
                    'record_count': len(table_data),
                    'data': table_data
                }
                total_records += len(table_data)
            
            # Add summary to metadata
            export_data['export_metadata']['total_tables'] = len(table_names)
            export_data['export_metadata']['total_records'] = total_records
            export_data['export_metadata']['tables_exported'] = table_names
            
            logger.info(f"Database export completed: {len(table_names)} tables, {total_records} records")
            return export_data
            
        except Exception as e:
            logger.error(f"Error exporting database: {e}")
            return None
    
    @staticmethod
    def export_specific_tables(table_list, decrypt_data=False):
        """Export specific tables to JSON format with optional decryption"""
        try:
            export_data = {
                'export_metadata': {
                    'export_timestamp': datetime.utcnow().isoformat(),
                    'database_name': 'monthly_organics',
                    'export_type': 'selective_tables',
                    'requested_tables': table_list,
                    'data_decrypted': decrypt_data
                },
                'tables': {}
            }
            
            total_records = 0
            exported_tables = []
            
            for table_name in table_list:
                logger.info(f"Exporting table: {table_name} (decrypt: {decrypt_data})")
                table_data = DatabaseExporter.get_table_data(table_name, decrypt_data=decrypt_data)
                
                if table_data is not None:  # Even empty tables should be included
                    export_data['tables'][table_name] = {
                        'record_count': len(table_data),
                        'data': table_data
                    }
                    total_records += len(table_data)
                    exported_tables.append(table_name)
            
            # Add summary to metadata
            export_data['export_metadata']['total_tables'] = len(exported_tables)
            export_data['export_metadata']['total_records'] = total_records
            export_data['export_metadata']['tables_exported'] = exported_tables
            
            logger.info(f"Selective export completed: {len(exported_tables)} tables, {total_records} records")
            return export_data
            
        except Exception as e:
            logger.error(f"Error exporting specific tables: {e}")
            return None
    
    @staticmethod
    def export_to_csv(tables_data, export_type='full'):
        """Export data to CSV format with separate files for each table"""
        try:
            import zipfile
            
            # Create a ZIP file containing CSV files for each table
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add metadata file
                metadata = {
                    'export_timestamp': datetime.utcnow().isoformat(),
                    'database_name': 'monthly_organics',
                    'export_type': export_type,
                    'format': 'csv'
                }
                
                if isinstance(tables_data, dict) and 'tables' in tables_data:
                    metadata.update({
                        'total_tables': tables_data['export_metadata']['total_tables'],
                        'total_records': tables_data['export_metadata']['total_records'],
                        'tables_exported': tables_data['export_metadata']['tables_exported']
                    })
                    tables = tables_data['tables']
                else:
                    tables = tables_data
                
                # Add metadata as JSON
                zip_file.writestr('export_metadata.json', json.dumps(metadata, indent=2))
                
                # Process each table
                for table_name, table_info in tables.items():
                    if isinstance(table_info, dict) and 'data' in table_info:
                        table_data = table_info['data']
                    else:
                        table_data = table_info
                    
                    if not table_data:
                        continue
                    
                    # Create CSV content for this table
                    csv_output = io.StringIO()
                    
                    if table_data:
                        # Get column headers from first row
                        headers = list(table_data[0].keys())
                        writer = csv.writer(csv_output)
                        writer.writerow(headers)
                        
                        # Write data rows
                        for row in table_data:
                            writer.writerow([str(row.get(header, '')) for header in headers])
                    
                    # Add CSV file to ZIP
                    zip_file.writestr(f'{table_name}.csv', csv_output.getvalue())
                    csv_output.close()
            
            zip_buffer.seek(0)
            return zip_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            return None
    
    @staticmethod
    def export_to_xlsx(tables_data, export_type='full'):
        """Export data to XLSX format with separate sheets for each table"""
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            import pandas as pd
            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)
            
            if isinstance(tables_data, dict) and 'tables' in tables_data:
                tables = tables_data['tables']
                metadata = tables_data['export_metadata']
            else:
                tables = tables_data
                metadata = {
                    'export_timestamp': datetime.utcnow().isoformat(),
                    'database_name': 'monthly_organics',
                    'export_type': export_type,
                    'format': 'xlsx'
                }
            
            # Add metadata sheet
            metadata_sheet = workbook.create_sheet('Export_Metadata')
            metadata_sheet['A1'] = 'Export Information'
            metadata_sheet['A1'].font = openpyxl.styles.Font(bold=True)
            
            row = 3
            for key, value in metadata.items():
                metadata_sheet[f'A{row}'] = str(key).replace('_', ' ').title()
                metadata_sheet[f'B{row}'] = str(value)
                row += 1
            
            # Process each table
            for table_name, table_info in tables.items():
                if isinstance(table_info, dict) and 'data' in table_info:
                    table_data = table_info['data']
                else:
                    table_data = table_info
                
                if not table_data:
                    continue
                
                # Get Excel-compatible data for this table
                # Check if metadata indicates decryption was used
                decrypt_data = metadata.get('data_decrypted', False)
                excel_table_data = DatabaseExporter.get_table_data(table_name, for_excel=True, decrypt_data=decrypt_data)
                
                # Create DataFrame from Excel-compatible table data
                df = pd.DataFrame(excel_table_data)
                
                # Create worksheet for this table
                # Excel sheet names can't exceed 31 characters
                sheet_name = table_name[:31] if len(table_name) > 31 else table_name
                worksheet = workbook.create_sheet(sheet_name)
                
                # Write DataFrame to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    worksheet.append(r)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Style header row
                for cell in worksheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Save to BytesIO
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to XLSX: {e}")
            return None
